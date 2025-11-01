# RIGa&DeepSeek 26.10.2025
from tkinter import ttk
import openpyxl
import csv
import os
from typing import Optional
from message_logger import MessageLogger


class ExcelUtils:
    """
    Utility class for Excel operations with integrated message logging

    Attributes:
        message_logger (MessageLogger): Logger for displaying messages in UI
    """

    def __init__(self, message_logger: MessageLogger):
        self.message_logger = message_logger
        self.encoding = 'cp1251'  # MS-DOS compatible encoding
        self.line_terminator = '\r\n'  # MS-DOS line endings

    def excel_to_csv_msdos(self, input_path: str, output_path: Optional[str] = None) -> Optional[str]:
        """
        Convert Excel file to MS-DOS CSV format with error logging

        Args:
            input_path: Path to input Excel file (.xlsx)
            output_path: Optional output CSV path (default: same as input with .csv extension)

        Returns:
            str: Path to created CSV file if successful, None otherwise
        """
        try:
            # Validate input
            if not os.path.exists(input_path):
                raise FileNotFoundError(f"Input file does not exist: {input_path}")

            if not input_path.lower().endswith('.xlsx'):
                raise ValueError("Only .xlsx files are supported")

            # Set default output path if not provided
            output_path = output_path or os.path.splitext(input_path)[0] + '.csv'

            # Load workbook
            wb = openpyxl.load_workbook(input_path, data_only=True)

            # Check sheet count
            if len(wb.sheetnames) > 1:
                raise ValueError("Excel file contains multiple sheets. Only single-sheet files are supported.")

            sheet = wb.active

            # Write CSV
            try:
                with open(output_path, 'w', newline='', encoding=self.encoding) as csvfile:
                    writer = csv.writer(
                        csvfile,
                        delimiter=',',
                        quoting=csv.QUOTE_MINIMAL,
                        lineterminator=self.line_terminator
                    )

                    for row in sheet.iter_rows():
                        row_values = []
                        for cell in row:
                            cell_value = str(cell.value) if cell.value is not None else ''
                            row_values.append(cell_value)
                        writer.writerow(row_values)

                self.message_logger.log_message('SUCCESS', f'File converted successfully: {output_path}')
                return output_path

            except UnicodeEncodeError:
                self.message_logger.log_message('ERROR', f'Original encoding failed')
                return None

            finally:
                wb.close()

        except Exception as e:
            self.message_logger.log_message('ERROR', f'Conversion failed: {str(e)}')
            return None

    def load_excel_to_treeview(self, excel_path: str, tree: ttk.Treeview, skip_rows: int = 1) -> bool:
        """
        Load data from Excel file into Treeview widget

        Args:
            excel_path: Path to Excel file
            tree: Treeview widget to populate
            skip_rows: Number of header rows to skip

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Load Excel data
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    cell_value = str(cell.value) if cell.value is not None else ''
                    row_values.append(cell_value)
                # Insert rows into Treeview
                tree.insert("", "end", values=row_values)

            workbook.close()
            return True

        except Exception as e:
            self.message_logger.log_message('ERROR', f"Error loading Excel: {str(e)}")
            return False

    def load_csv_to_treeview(self, csv_path: str, tree: ttk.Treeview, skip_rows: int = 0, delimiter: str = ',') -> bool:
        """
        Load data from CSV file into Treeview widget with dynamic columns

        Args:
            csv_path: Path to CSV file
            tree: Treeview widget to populate
            skip_rows: Number of header rows to skip
            delimiter: CSV delimiter character

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read CSV file with different encodings
            encodings = ['utf-8', 'cp1251', 'latin-1']
            data_rows = []

            for encoding in encodings:
                try:
                    with open(csv_path, 'r', encoding=encoding) as csv_file:
                        csv_reader = csv.reader(csv_file, delimiter=delimiter)
                        data_rows = list(csv_reader)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                self.message_logger.log_message('ERROR', "Failed to read CSV file with any encoding")
                return False

            if not data_rows:
                self.message_logger.log_message('WARNING', "CSV file is empty")
                return True

            # Use first row as column headers
            headers = data_rows[0] if data_rows else []

            # Skip header row for data
            data_rows = data_rows[1:] if len(data_rows) > 1 else []

            if not headers:
                self.message_logger.log_message('WARNING', "No headers found in CSV file")
                return True

            # Configure Treeview columns
            tree["columns"] = [f"col_{i}" for i in range(len(headers))]

            # Set column headings from CSV first row
            for i, header in enumerate(headers):
                tree.heading(f"col_{i}", text=str(header))
                tree.column(f"col_{i}", width=100)

            # Insert all data rows
            for row in data_rows:
                # Pad row with empty values if it has fewer columns than headers
                padded_row = row + [''] * (len(headers) - len(row))
                tree.insert("", "end", values=padded_row)

            return True

        except Exception as e:
            self.message_logger.log_message('ERROR', f"Error loading CSV: {str(e)}")
            return False

    def save_list_to_csv(self, data: list, csv_path: str):
        try:
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(
                    csvfile,
                    delimiter=',',
                    quoting=csv.QUOTE_MINIMAL,
                    lineterminator='\r\n'
                )

                for row in data:
                    writer.writerow(row)

            self.message_logger.log_message('SUCCESS', f"CSV file saved: {csv_path}")
            return True

        except Exception as e:
            self.message_logger.log_message('ERROR', f"Failed to save CSV: {str(e)}")
            return False

    def open_or_create_xlsx(self, xlsx_path):
        """Open file in Microsoft Excel. Creates the file if it doesn't exist."""
        try:
            if not os.path.exists(xlsx_path):
                import openpyxl
                wb = openpyxl.Workbook()
                wb.save(xlsx_path)
                self.message_logger.log_message('SUCCESS', f"Created new Excel file: {xlsx_path}")

            os.startfile(xlsx_path)
            self.message_logger.log_message('SUCCESS', f"Opened Excel file: {xlsx_path}")

        except PermissionError:
            self.message_logger.log_message('ERROR', f"Permission denied when accessing: {xlsx_path}")
        except Exception as e:
            self.message_logger.log_message('ERROR', f"Failed to create/open Excel file: {str(e)}")