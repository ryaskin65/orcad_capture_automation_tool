# RIGa&DeepSeek 21.12.2025

import os
from typing import Dict, List, Tuple, Optional, Set, Any
from openpyxl import Workbook, load_workbook

from page_geometry import PageGeometry
from wire_data import WireData, WireDataProcessor
from offset_calculator import OffsetCalculator
from excel_utils import ExcelUtils


class DataConverter:
    """Main converter class for transforming cable data from input to output format"""

    def __init__(self, message_logger):
        self.message_logger = message_logger
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.page_geometry = PageGeometry()

        # Initialize helper classes
        self.wire_processor = WireDataProcessor(message_logger)
        self.offset_calculator = OffsetCalculator(message_logger)
        self.excel_utils = ExcelUtils(message_logger)

        # Initialize cache for wire calculation data
        self.page_wire_data = {}  # page_name -> List[wire_data_dict]

        # Define header fields that appear before PAGE sections
        self.header_fields = [
            "ProjectNumber",
            "NumberCable",
            "NameLeftSide",
            "NameRightSide",
            "Title",
            "PartNumber",
            "DocumentNumber",
            "Revision",
        ]

        # Define output column structure (fixed format)
        self.output_columns = [
            "Signal name",
            "Right connector",
            "Left connector",
            "Right side IN/OUT",
            "Type",
            "Left wire width, Gauge",
            "Color",
            "Right wire width, Gauge",
            "Right wire offset",  # Y offset for right side
            "Right group",  # X offset/group for right side (0,1,2 or "")
            "Left wire offset",  # Y offset for left side
            "Left group",  # X offset/group for left side (0,1,2 or "")
        ]

    def convert_excel_file(
        self, input_path: str, output_path: str
    ) -> Tuple[bool, List, List]:
        """
        Convert Excel file from input format to output format with full validation

        Process:
        1. Parse input Excel file
        2. Validate data structure
        3. Calculate wire coordinates, offsets, and groups
        4. Generate output file

        Returns: (success, errors, warnings)
        """
        self.errors = []
        self.warnings = []

        if not os.path.exists(input_path):
            self.errors.append(f"Input file does not exist: {input_path}")
            return False, self.errors, self.warnings

        try:
            # Load input workbook
            input_wb = load_workbook(input_path)
            input_ws = input_wb.active

            # Parse input data into structured format
            input_data = self._parse_input_worksheet(input_ws)
            if not input_data:
                self.errors.append("No valid data found in input file")
                return False, self.errors, self.warnings

            # Validate data structure and wiring rules
            is_valid, validation_errors, validation_warnings = (
                self._validate_data_structure(input_data)
            )
            self.errors.extend(validation_errors)
            self.warnings.extend(validation_warnings)

            if not is_valid:
                return False, self.errors, self.warnings

            # Calculate all wire data (coordinates, offsets, groups)
            success = self._calculate_all_wire_data(input_data)
            if not success:
                return False, self.errors, self.warnings

            # Convert data format and generate output structure
            success, converted_data = self._convert_and_process_data(input_data)
            if not success:
                return False, self.errors, self.warnings

            # Create output workbook
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.title = "Cable Data"

            # Write converted data to output worksheet
            self._write_output_worksheet(output_ws, converted_data)

            # Save output file
            output_wb.save(output_path)
            self.message_logger.log_message(
                "SUCCESS", f"File converted successfully: {output_path}"
            )

            return True, self.errors, self.warnings

        except Exception as e:
            self.errors.append(f"Conversion error: {str(e)}")
            return False, self.errors, self.warnings

    def _parse_input_worksheet(self, worksheet) -> Dict:
        """
        Parse input worksheet and extract structured data

        Returns: Dictionary with structure:
        {
            'headers': {
                'ProjectNumber': value,
                'Title': value,
                ...
            },
            'pages': {
                'page_name': {
                    'wire_headers': [column names],
                    'wires': [WireData objects]
                }
            }
        }
        """
        data = {"headers": {}, "pages": {}}

        current_section = "headers"
        current_page = None
        column_mapping = {}
        header_found = False

        for row in worksheet.iter_rows(values_only=True):
            # Skip empty rows
            if not row or not any(cell for cell in row if cell and str(cell).strip()):
                continue

            first_cell = str(row[0]).strip() if row[0] else ""

            # Detect PAGE section - marks start of a cable page
            if first_cell.upper() == "PAGE":
                current_section = "pages"
                if len(row) > 1 and row[1]:
                    current_page = str(row[1]).strip()
                    data["pages"][current_page] = {"wire_headers": [], "wires": []}
                    header_found = False
                    column_mapping = {}
                continue

            # Process header section (before first PAGE)
            if current_section == "headers":
                if first_cell in self.header_fields and len(row) > 1:
                    data["headers"][first_cell] = str(row[1]) if row[1] else ""

            # Process page section (after PAGE marker)
            elif current_section == "pages" and current_page:
                # Detect wire headers row (starts with "Signal name")
                if first_cell == "Signal name" and not header_found:
                    header_found = True
                    data["pages"][current_page]["wire_headers"] = [
                        str(cell).strip() if cell else "" for cell in row
                    ]

                    # Create column mapping for this page
                    column_mapping = self._create_column_mapping(
                        data["pages"][current_page]["wire_headers"]
                    )
                    if not column_mapping:
                        self.errors.append(
                            f"Page '{current_page}': Required columns not found"
                        )

                # Process wire data rows (after headers, not header fields)
                elif (
                    header_found and first_cell and first_cell not in self.header_fields
                ):
                    wire_data = self.wire_processor.parse_wire_data(row, column_mapping)
                    if wire_data:
                        data["pages"][current_page]["wires"].append(wire_data)

        return data

    def _create_column_mapping(self, headers: List[str]) -> Dict[str, int]:
        """
        Create mapping from column names to indices

        Handles flexible input format where columns may be in different positions
        """
        mapping = {}

        for i, header in enumerate(headers):
            if not header:
                continue

            exact_header = str(header).strip()
            clean_header = exact_header.lower()

            # Map column names with different spellings/cases to standard names
            column_mappings = {
                "signal name": "Signal name",
                "right connector": "Right connector",
                "left connector": "Left connector",
                "left side in/out": "Left side IN/OUT",
                "type": "Type",
                "left wire width, gauge": "Left wire width, Gauge",
                "color": "Color",
                "right wire width, gauge": "Right wire width, Gauge",
                "connect to right wire with number": "Connect to right wire with number",
                "connect to left wire with number": "Connect to left wire with number",
                "left wire width gauge": "Left wire width, Gauge",
                "right wire width gauge": "Right wire width, Gauge",
            }

            if clean_header in column_mappings:
                mapping[column_mappings[clean_header]] = i

        return mapping

    def _validate_data_structure(self, data: Dict) -> Tuple[bool, List, List]:
        """
        Validate cable data structure and wiring rules

        Returns: (is_valid, errors, warnings)
        """
        errors = []
        warnings = []

        if not data.get("pages"):
            errors.append("No pages found in cable data")
            return False, errors, warnings

        # Required columns in input data
        required_headers = [
            "Signal name",
            "Right connector",
            "Left connector",
        ]

        for page_name, page_data in data["pages"].items():
            headers = page_data.get("wire_headers", [])
            wires = page_data.get("wires", [])

            # 1. Basic header validation
            # Check required headers exist
            for req_header in required_headers:
                if req_header not in headers:
                    errors.append(
                        f"Page '{page_name}': Missing required header '{req_header}'"
                    )

            # 2. Wire type rules (twisted pairs)
            # Check wire data exists
            if not wires:
                warnings.append(f"Page '{page_name}': No wire data found")
            # Validate wire type rules (twisted pairs must come in pairs)
            wire_errors, wire_warnings = self.wire_processor.validate_wire_type_rules(
                page_name, wires
            )
            errors.extend(wire_errors)
            warnings.extend(wire_warnings)

            # 3. Connector pin formats and uniqueness
            # Validate connector pin formats and uniqueness
            pin_errors, pin_warnings = self.wire_processor.validate_connector_pins(
                page_name, wires
            )
            errors.extend(pin_errors)
            warnings.extend(pin_warnings)

            # 4. Signal name uniqueness
            # Validate signal name uniqueness
            signal_errors, signal_warnings = self.wire_processor.validate_signal_names(
                page_name, wires
            )
            errors.extend(signal_errors)
            warnings.extend(signal_warnings)

            # 5. Splice offset rules
            # Validate splice rules
            splice_errors, splice_warnings = self.wire_processor.validate_splice_rules(
                page_name, wires
            )
            errors.extend(splice_errors)
            warnings.extend(splice_warnings)

            # 6. Group count rules (warning only)
            # Validate group count rules (warning only, full check in OffsetCalculator)
            group_errors, group_warnings = self.wire_processor.validate_group_count_rules(
                page_name, wires
            )
            errors.extend(group_errors)
            warnings.extend(group_warnings)

            # 7. Connection consistency (no chains > 2, no cycles)
            # Validate connection consistency
            connection_errors, connection_warnings = (
                self.wire_processor.validate_connection_consistency(page_name, wires)
            )
            errors.extend(connection_errors)
            warnings.extend(connection_warnings)

        return len(errors) == 0, errors, warnings

    def _calculate_all_wire_data(self, input_data: Dict) -> bool:
        """
        Calculate all wire data and validate group count
        """
        for page_name, page_data in input_data["pages"].items():
            wires = page_data.get("wires", [])

            if not wires:
                continue

            try:
                wire_data_array = self.offset_calculator.calculate_wire_data(wires)
                self.page_wire_data[page_name] = wire_data_array

                # Page capacity: reject layouts that would overflow the sheet
                # (mirrors cable.tcl's connector-rectangle bounds check).
                fits, capacity_msg, _info = self.page_geometry.validate_page(
                    wire_data_array
                )
                if not fits:
                    self.errors.append(f"Page '{page_name}': {capacity_msg}")
                    return False

                # NEW: Check group count validation
                left_groups_used = set()
                right_groups_used = set()

                for data in wire_data_array:
                    left_group = data["left_x_offset"]
                    right_group = data["right_x_offset"]

                    if left_group:
                        left_groups_used.add(left_group)
                    if right_group:
                        right_groups_used.add(right_group)

                # Remove empty strings (group 0 when not written)
                left_groups_used.discard("")
                right_groups_used.discard("")

                # Check if we have more than 2 groups (0,1)
                if len(left_groups_used) > 2:
                    self.errors.append(
                        f"Page '{page_name}': Required {len(left_groups_used)} X-offset groups on left side. "
                        f"Maximum allowed is 2. Splices with intersecting Y-ranges need different X-offsets."
                    )
                    return False

                if len(right_groups_used) > 2:
                    self.errors.append(
                        f"Page '{page_name}': Required {len(right_groups_used)} X-offset groups on right side. "
                        f"Maximum allowed is 2. Splices with intersecting Y-ranges need different X-offsets."
                    )
                    return False

            except Exception as e:
                self.errors.append(
                    f"Error calculating wire data for page '{page_name}': {str(e)}"
                )
                return False

        return True

    def _convert_and_process_data(self, input_data: Dict) -> Tuple[bool, List[List]]:
        """
        Convert data format and generate output structure

        Returns: (success, converted_data_list)
        """
        converted_data = []

        # Add header rows from input data
        for header_field in self.header_fields:
            if header_field in input_data["headers"]:
                converted_data.append(
                    [header_field, input_data["headers"][header_field]]
                )

        # Add empty row before pages
        converted_data.append([])

        # Process each page
        for page_name, page_data in input_data["pages"].items():
            # Add PAGE marker row
            converted_data.append(["PAGE", page_name])

            # Add output column headers
            converted_data.append(self.output_columns)

            # Get wires and calculated data for this page
            wires = page_data.get("wires", [])
            wire_data_array = self.page_wire_data.get(page_name, [])

            if len(wire_data_array) != len(wires):
                self.errors.append(
                    f"Page '{page_name}': Wire calculation data mismatch"
                )
                return False, []

            # Process each wire
            for i, wire in enumerate(wires):
                wire_calc_data = wire_data_array[i]

                # Format output row with all calculated values
                output_row = self._format_output_row(wire, wire_calc_data)
                converted_data.append(output_row)

            # Add empty row between pages
            converted_data.append([])

        return True, converted_data

    def _format_output_row(
        self, wire: WireData, wire_calc_data: Dict[str, Any]
    ) -> List[str]:
        """
        Format a single wire's data into output row format

        FIXED: Only set offset=0 for ACTUAL splice wires
        """
        # Convert IN/OUT direction
        right_in_out = self._convert_in_out_direction(wire)

        # Get calculated offsets and groups
        right_y_offset = wire_calc_data.get("right_y_offset", "")
        right_group = wire_calc_data.get("right_x_offset", "")
        left_y_offset = wire_calc_data.get("left_y_offset", "")
        left_group = wire_calc_data.get("left_x_offset", "")

        # IMPORTANT: Clear offset if wire is not actually a splice on that side
        is_splice = wire_calc_data.get("is_splice", False)

        # Only set offset=0 if wire is ACTUALLY a target on that side
        if right_y_offset == "0" and not wire_calc_data.get("is_right_target", False):
            right_y_offset = ""  # Not a right splice

        if left_y_offset == "0" and not wire_calc_data.get("is_left_target", False):
            left_y_offset = ""  # Not a left splice

        return [
            wire.signal_name,
            wire.right_connector,
            wire.left_connector,
            right_in_out,
            wire.wire_type,
            wire.left_gauge,
            wire.color,
            wire.right_gauge,
            right_y_offset,
            right_group,
            left_y_offset,
            left_group,
        ]

    def _convert_in_out_direction(self, wire: WireData) -> str:
        """
        Convert Left side IN/OUT to Right side IN/OUT

        Rule: IN ↔ OUT (invert direction)
        """
        left_in_out = wire.left_in_out.upper() if wire.left_in_out else ""

        if left_in_out == "IN":
            return "OUT"
        elif left_in_out == "OUT":
            return "IN"
        else:
            return left_in_out  # Return unchanged if not IN/OUT

    def _write_output_worksheet(self, worksheet, converted_data: List[List]):
        """
        Write converted data to output worksheet

        Args:
            worksheet: OpenPyXL worksheet object
            converted_data: 2D list of data to write
        """
        for row_idx, row_data in enumerate(converted_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def validate_input_data(self, input_path: str) -> Tuple[bool, List[str], List[str]]:
        """
        Validate input Excel file without converting it
        """
        self.errors = []
        self.warnings = []

        try:
            input_wb = load_workbook(input_path)
            input_ws = input_wb.active

            # Parse input data
            input_data = self._parse_input_worksheet(input_ws)
            if not input_data:
                self.errors.append("No valid data found in input file")
                return False, self.errors, self.warnings

            # Validate ALL rules including new splice rules
            is_valid, validation_errors, validation_warnings = (
                self._validate_data_structure(input_data)
            )
            self.errors.extend(validation_errors)
            self.warnings.extend(validation_warnings)

            return is_valid, self.errors, self.warnings

        except Exception as e:
            self.errors.append(f"Validation error: {str(e)}")
            return False, self.errors, self.warnings

    def get_conversion_summary(self) -> str:
        """
        Get formatted conversion summary for display

        Returns: String with errors, warnings, or success message
        """
        summary = []

        if self.errors:
            summary.append("CONVERSION ERRORS:")
            for error in self.errors:
                summary.append(f"  ❌ {error}")

        if self.warnings:
            summary.append("CONVERSION WARNINGS:")
            for warning in self.warnings:
                summary.append(f"  ⚠️  {warning}")

        if not self.errors and not self.warnings:
            summary.append("✅ File conversion completed successfully")
            summary.append(
                "📝 Data validated and auto-processed with offsets and groups"
            )

        return "\n".join(summary)
