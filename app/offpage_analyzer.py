# 16.01.2026 RIGa&AI
# offpage_analyzer.py
"""
OffPage Analyzer - analyzes offpage connector data from CSV files.
Generates reports about connector names, frequencies, and spatial relationships.
"""

import csv
import os
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from message_logger import MessageLogger


class OffPageAnalyzer:
    """
    Analyzes offpage connector data from CSV files and generates Excel reports.
    Provides analysis of:
    1. All_Names - all connection names with total count and pages
    2. Frequent_Names - names appearing more than twice on any page
    3. Opposite_Names - names opposite each other on the same page
    """

    def __init__(self, message_logger: MessageLogger):
        """
        Initialize OffPageAnalyzer with message logger.

        Args:
            message_logger: MessageLogger instance for logging messages
        """
        self.message_logger = message_logger
        self.page_data = defaultdict(list)  # page_name -> list of (x, y, name)
        self.name_page_counts = defaultdict(
            lambda: defaultdict(int)
        )  # name -> {page: count}
        self.name_total_counts = Counter()  # name -> total count across all pages
        self.all_names = []  # List of (name, connector_name, total_count, pages)
        self.frequent_names = []  # List of (name, connector_name, total_count, pages)
        self.opposite_pairs = (
            []
        )  # List of (page, name1, loc1, name2, loc2, dist_x, dist_y)
        self.total_pages = 0

    def extract_connector_name(self, name: str) -> str:
        """
        Extract connector name from full connection name based on rules:
        1. If '/' exists -> part before '/' is connector name
        2. Else if '-' exists -> part before '-' is connector name
        3. Else if '_' exists -> part before '_' is connector name
        4. Else -> empty string (no connector name)

        Args:
            name: Full connection name

        Returns:
            Connector name or empty string
        """
        if "/" in name:
            return name.split("/")[0]
        elif "-" in name:
            return name.split("-")[0]
        elif "_" in name:
            return name.split("_")[0]
        else:
            return ""

    def parse_csv_data(self, content: str):
        """
        Parse CSV content and extract page data.

        Args:
            content: CSV file content as string

        Returns:
            List of (page_name, page_data) tuples
        """
        pages = []
        current_page = None
        current_data = []

        lines = content.split("\n")

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Start of new page
            if line.startswith(">>> PAGE:"):
                if current_page is not None:
                    pages.append((current_page, current_data))
                current_page = line.replace(">>> PAGE: ", "").strip()
                current_data = []
                continue

            # Skip header row
            if line.startswith("X") and "," in line:
                continue

            # Data row
            if current_page is not None and line:
                # Split by comma
                parts = line.split(",")
                i = 0
                while i < len(parts) - 2:
                    # Check if we have a valid triple: X, Y, Name
                    x_str = parts[i].strip()
                    y_str = parts[i + 1].strip()
                    name = parts[i + 2].strip()

                    # Skip if name is empty or if x/y are not numbers
                    if name and x_str and y_str:
                        try:
                            x = int(x_str)
                            y = int(y_str)
                            # Add to current page data
                            current_data.append((x, y, name))
                            # Move forward by 3 (X, Y, Name)
                            i += 3
                            # Skip empty fields between triples
                            while i < len(parts) and not parts[i].strip():
                                i += 1
                            continue
                        except ValueError:
                            # If conversion fails, move to next
                            i += 1
                    else:
                        i += 1

        if current_page is not None:
            pages.append((current_page, current_data))

        return pages

    def analyze_csv_file(self, csv_path: str):
        """
        Analyze CSV file and collect statistics.

        Args:
            csv_path: Path to CSV file to analyze

        Returns:
            bool: True if successful, False otherwise
        """
        if not os.path.exists(csv_path):
            self.message_logger.log_message("ERROR", f"CSV file not found: {csv_path}")
            return False

        try:
            # Read CSV file with different encodings
            encodings = ["utf-8", "cp1251", "latin-1", "utf-8-sig"]
            content = None

            for encoding in encodings:
                try:
                    with open(csv_path, "r", encoding=encoding) as f:
                        content = f.read()
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:
                self.message_logger.log_message(
                    "ERROR", "Failed to read CSV file with any encoding"
                )
                return False

            # Parse data
            pages = self.parse_csv_data(content)
            self.total_pages = len(pages)

            # Clear previous data
            self.page_data.clear()
            self.name_page_counts.clear()
            self.name_total_counts.clear()
            self.all_names.clear()
            self.frequent_names.clear()
            self.opposite_pairs.clear()

            # Collect data from all pages
            for page_name, connections in pages:
                page_items = []

                for x, y, name in connections:
                    # Store connection in page data
                    page_items.append((x, y, name))

                    # Count name occurrences on this page
                    self.name_page_counts[name][page_name] += 1

                    # Count total occurrences across all pages
                    self.name_total_counts[name] += 1

                # Store page connections for opposite name detection
                self.page_data[page_name] = page_items

            # Collect all names (each name appears once in the list)
            self.all_names = []
            for name in sorted(self.name_page_counts.keys()):
                total_count = self.name_total_counts[name]
                connector_name = self.extract_connector_name(name)

                # Get pages where this name appears
                page_counts = self.name_page_counts[name]
                pages_list = []

                for page, count in sorted(page_counts.items()):
                    # Store page name and count
                    pages_list.append({"page": page, "count": count})

                self.all_names.append(
                    {
                        "name": name,
                        "connector_name": connector_name,
                        "total_count": total_count,
                        "pages": pages_list,
                    }
                )

            # Analyze frequent names (names that appear more than twice on any page)
            self._analyze_frequent_names()

            # Find opposite names on each page
            self._find_opposite_names()

            return True

        except Exception as e:
            self.message_logger.log_message("ERROR", f"Analysis failed: {str(e)}")
            return False

    def _analyze_frequent_names(self):
        """
        Analyze and collect names that appear more than twice on any page.
        For each name, collect pages where count > 2.
        """
        frequent_names_data = []

        for name, page_counts in self.name_page_counts.items():
            # Get pages where this name appears more than twice
            frequent_pages = []
            for page, count in sorted(page_counts.items()):
                if count >= 3:  # Changed from >2 to >=3 as requested
                    frequent_pages.append({"page": page, "count": count})

            if frequent_pages:
                total_count = self.name_total_counts[name]
                connector_name = self.extract_connector_name(name)

                frequent_names_data.append(
                    {
                        "name": name,
                        "connector_name": connector_name,
                        "total_count": total_count,
                        "pages": frequent_pages,
                    }
                )

        # Sort by total count (descending), then by name
        self.frequent_names = sorted(
            frequent_names_data, key=lambda x: (-x["total_count"], x["name"])
        )

    def _find_opposite_names(self):
        """
        Find names that are opposite each other on the same page.
        Conditions: ΔX ≤ 200 and ΔY ≤ 10, and names are different.
        """
        for page_name, connections in self.page_data.items():
            if len(connections) < 2:
                continue

            # Check all pairs of connections on this page
            for i in range(len(connections)):
                x1, y1, name1 = connections[i]

                for j in range(i + 1, len(connections)):
                    x2, y2, name2 = connections[j]

                    # Skip same name
                    if name1 == name2:
                        continue

                    # Calculate distances
                    dist_x = abs(x1 - x2)
                    dist_y = abs(y1 - y2)

                    # Check if opposite each other
                    if dist_x <= 200 and dist_y <= 10:
                        self.opposite_pairs.append(
                            {
                                "page": page_name,
                                "name1": name1,
                                "loc1": (x1, y1),
                                "name2": name2,
                                "loc2": (x2, y2),
                                "dist_x": dist_x,
                                "dist_y": dist_y,
                            }
                        )

    def generate_excel_report(self, output_path: str):
        """
        Generate Excel report with analysis results.

        Args:
            output_path: Path to save Excel report

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            wb = Workbook()

            # Style definitions
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            title_font = Font(size=14, bold=True)
            bold_font = Font(bold=True)

            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Title
            ws_summary["A1"] = "OffPage Connector Analysis Report"
            ws_summary["A1"].font = title_font
            ws_summary.merge_cells("A1:E1")
            ws_summary["A1"].alignment = Alignment(horizontal="center")

            # Summary statistics
            ws_summary["A3"] = "Analysis Statistics"
            ws_summary["A3"].font = Font(size=12, bold=True, color="000080")

            summary_data = [
                ["Total Pages Analyzed", self.total_pages],
                ["Total Connection Names", len(self.all_names)],
                ["Names with ≥3 occurrences on page", len(self.frequent_names)],
                ["Opposite Pairs Found (ΔX≤200, ΔY≤10)", len(self.opposite_pairs)],
            ]

            for i, (label, value) in enumerate(summary_data, start=4):
                ws_summary[f"A{i}"] = label
                ws_summary[f"A{i}"].font = bold_font
                ws_summary[f"B{i}"] = value

            # Sheet 2: All_Names
            ws_all_names = wb.create_sheet("All_Names")
            ws_all_names.title = "All_Names"

            # Title
            ws_all_names["A1"] = "All Connection Names"
            ws_all_names["A1"].font = title_font
            ws_all_names.merge_cells("A1:D1")
            ws_all_names["A1"].alignment = Alignment(horizontal="center")

            # Headers
            headers = [
                "Connection Name",
                "Connector Name",
                "Total Count",
                "Page 1",
                "Page 2",
                "Page 3",
                "Page 4",
                "Page 5",
                "Page 6",
                "Page 7",
                "Page 8",
                "Page 9",
                "Page 10",
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws_all_names.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows for All_Names
            row = 4
            for item in self.all_names:
                name = item["name"]
                connector_name = item["connector_name"]
                total_count = item["total_count"]
                pages = item["pages"]

                # Column A: Connection Name
                ws_all_names[f"A{row}"] = name

                # Column B: Connector Name
                ws_all_names[f"B{row}"] = connector_name

                # Column C: Total Count
                ws_all_names[f"C{row}"] = total_count

                # Columns D onward: Pages (just page names)
                for i, page_info in enumerate(pages, start=4):
                    if i <= len(headers):  # Stay within header columns
                        ws_all_names.cell(row=row, column=i).value = page_info["page"]

                row += 1

            # Sheet 3: Frequent_Names
            ws_frequent = wb.create_sheet("Frequent_Names")
            ws_frequent.title = "Frequent_Names"

            if self.frequent_names:
                # Title
                ws_frequent["A1"] = "Frequent Names (≥3 occurrences on page)"
                ws_frequent["A1"].font = title_font
                ws_frequent.merge_cells("A1:D1")
                ws_frequent["A1"].alignment = Alignment(horizontal="center")

                # Headers (same as All_Names)
                headers = [
                    "Connection Name",
                    "Connector Name",
                    "Total Count",
                    "Page 1",
                    "Page 2",
                    "Page 3",
                    "Page 4",
                    "Page 5",
                    "Page 6",
                    "Page 7",
                    "Page 8",
                    "Page 9",
                    "Page 10",
                ]

                for col, header in enumerate(headers, start=1):
                    cell = ws_frequent.cell(row=3, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Data rows for Frequent_Names
                row = 4
                for item in self.frequent_names:
                    name = item["name"]
                    connector_name = item["connector_name"]
                    total_count = item["total_count"]
                    pages = item["pages"]

                    # Column A: Connection Name
                    ws_frequent[f"A{row}"] = name

                    # Column B: Connector Name
                    ws_frequent[f"B{row}"] = connector_name

                    # Column C: Total Count
                    ws_frequent[f"C{row}"] = total_count

                    # Columns D onward: Pages with counts
                    for i, page_info in enumerate(pages, start=4):
                        if i <= len(headers):  # Stay within header columns
                            # Format: "PageName (Count)"
                            cell_value = f"{page_info['page']} ({page_info['count']})"
                            ws_frequent.cell(row=row, column=i).value = cell_value

                    row += 1
            else:
                ws_frequent["A1"] = "No Frequent Names Found"
                ws_frequent["A1"].font = title_font
                ws_frequent["A1"].alignment = Alignment(horizontal="center")

            # Sheet 4: Opposite_Names
            ws_opposite = wb.create_sheet("Opposite_Names")
            ws_opposite.title = "Opposite_Names"

            if self.opposite_pairs:
                # Title
                ws_opposite["A1"] = "Opposite Connection Names (ΔX ≤ 200, ΔY ≤ 10)"
                ws_opposite["A1"].font = title_font
                ws_opposite.merge_cells("A1:G1")
                ws_opposite["A1"].alignment = Alignment(horizontal="center")

                # Headers
                headers = [
                    "Page",
                    "Name 1",
                    "Location 1 (X,Y)",
                    "Name 2",
                    "Location 2 (X,Y)",
                    "ΔX",
                    "ΔY",
                ]
                for col, header in enumerate(headers, start=1):
                    cell = ws_opposite.cell(row=3, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Data rows
                row = 4
                # Group by page for better readability
                opposite_by_page = defaultdict(list)
                for pair in self.opposite_pairs:
                    opposite_by_page[pair["page"]].append(pair)

                for page_name in sorted(opposite_by_page.keys()):
                    # Page header
                    ws_opposite[f"A{row}"] = f"Page: {page_name}"
                    ws_opposite[f"A{row}"].font = Font(bold=True, color="FF0000")
                    ws_opposite.merge_cells(f"A{row}:G{row}")
                    ws_opposite[f"A{row}"].alignment = Alignment(horizontal="center")
                    row += 1

                    # Pairs for this page
                    for pair in opposite_by_page[page_name]:
                        ws_opposite[f"A{row}"] = page_name
                        ws_opposite[f"B{row}"] = pair["name1"]
                        ws_opposite[f"C{row}"] = (
                            f"({pair['loc1'][0]}, {pair['loc1'][1]})"
                        )
                        ws_opposite[f"D{row}"] = pair["name2"]
                        ws_opposite[f"E{row}"] = (
                            f"({pair['loc2'][0]}, {pair['loc2'][1]})"
                        )
                        ws_opposite[f"F{row}"] = pair["dist_x"]
                        ws_opposite[f"G{row}"] = pair["dist_y"]
                        row += 1

                    # Empty row between pages
                    row += 1
            else:
                ws_opposite["A1"] = "No Opposite Connection Names Found"
                ws_opposite["A1"].font = title_font
                ws_opposite["A1"].alignment = Alignment(horizontal="center")

            # Adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Add auto-filter for All_Names and Frequent_Names sheets
            for ws in [ws_all_names, ws_frequent]:
                if ws.max_row > 3:
                    last_col = get_column_letter(len(headers))
                    ws.auto_filter.ref = f"A3:{last_col}3"

            # Freeze panes for headers
            for ws in [ws_all_names, ws_frequent, ws_opposite]:
                if ws.max_row > 3:
                    ws.freeze_panes = "A4"

            # Save workbook
            wb.save(output_path)
            return True

        except Exception as e:
            self.message_logger.log_message(
                "ERROR", f"Failed to generate Excel report: {str(e)}"
            )
            return False

    def analyze_and_generate_report(self, csv_path: str, report_path: str = None):
        """
        Complete analysis workflow: analyze CSV and generate Excel report.

        Args:
            csv_path: Path to input CSV file
            report_path: Path for output Excel report (optional)

        Returns:
            bool: True if successful, False otherwise
        """
        if report_path is None:
            report_path = os.path.join(
                os.path.dirname(csv_path), "offpage_analysis_report.xlsx"
            )

        self.message_logger.log_message("INFO", f"Starting analysis of: {csv_path}")

        # Analyze CSV file
        if not self.analyze_csv_file(csv_path):
            return False

        # Log analysis statistics
        self.message_logger.log_message("INFO", "Analysis completed successfully:")
        self.message_logger.log_message(
            "INFO", f"  - Total pages analyzed: {self.total_pages}"
        )
        self.message_logger.log_message(
            "INFO", f"  - Total connection names found: {len(self.all_names)}"
        )

        if self.frequent_names:
            self.message_logger.log_message(
                "INFO",
                f"  - Names with ≥3 occurrences on page: {len(self.frequent_names)}",
            )
        else:
            self.message_logger.log_message(
                "INFO", "  - No names with ≥3 occurrences on page"
            )

        if self.opposite_pairs:
            self.message_logger.log_message(
                "INFO",
                f"  - Opposite connection pairs found: {len(self.opposite_pairs)}",
            )
        else:
            self.message_logger.log_message(
                "INFO", "  - No opposite connection pairs found"
            )

        # Generate Excel report
        if not self.generate_excel_report(report_path):
            return False

        self.message_logger.log_message(
            "SUCCESS", f"Excel report saved to: {report_path}"
        )
        return True
